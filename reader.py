from unidecode import unidecode
import xlsxwriter
import openpyxl
import os

all_data = []

facnum = 1

def printProgressBar (iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█', printEnd = "\r"):
    """
    Call in a loop to create terminal progress bar
    @params:
        iteration   - Required  : current iteration (Int)
        total       - Required  : total iterations (Int)
        prefix      - Optional  : prefix string (Str)
        suffix      - Optional  : suffix string (Str)
        decimals    - Optional  : positive number of decimals in percent complete (Int)
        length      - Optional  : character length of bar (Int)
        fill        - Optional  : bar fill character (Str)
        printEnd    - Optional  : end character (e.g. "\r", "\r\n") (Str)
    """
    percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
    filledLength = int(length * iteration // total)
    bar = fill * filledLength + '-' * (length - filledLength)
    print(f'\r{prefix} |{bar}| {percent}% {suffix}', end = printEnd)
    # Print New Line on Complete
    if iteration == total: 
        print()

def read(loc, factor, n):
    global facnum
    wb_obj = openpyxl.load_workbook(loc)
    sheet = wb_obj.active
    data = []

    rows = list(sheet.rows)
    factor_num = None
    if rows[0][14].value != None:
        factor_num = (int(unidecode(str(rows[0][14].value).split("-")[0])))
    rows = rows[5:]

    if factor_num in factor:
        facnum +=1
        for row in rows:
            r = [x.value if x.value != None else '' for x in row]
            del r[0]
            r.append(facnum)
            data.append(r)
            
        return data
    else:
        return None

def read_main(loc):
    wb_obj = openpyxl.load_workbook(loc)
    sheet = wb_obj.active
    data = []

    cols = list(sheet.columns)[0]

    for row in cols:
        if row.value != None:
            data.append(row.value)
        
    return data

home = os.path.expanduser('~')
dir = os.path.join(home, 'Documents/files/')
file_dir = os.path.join(dir, "Factors/")

files = os.listdir(file_dir)

factor_nums = read_main(dir + 'main.xlsx')

all_lenght = len(files)
n = 1

printProgressBar(0, all_lenght, prefix = 'Progress:', suffix = 'Complete', length = 50)

for i in files:
    if i.endswith(".xlsx"):
        # print(n, "/", all_lenght)
        printProgressBar(n + 1, all_lenght, prefix = 'Progress:', suffix = 'Complete', length = 50)
        n += 1
        try:
            d = read(os.path.join(file_dir, i), factor_nums, n)
            if d != None:
                all_data.append(d)
        except:
            pass


workbook = xlsxwriter.Workbook(os.path.join(dir, 'all.xlsx'))
worksheet = workbook.add_worksheet()

row = 0
column = 0

for i in all_data:
    for j in i:
        column = 0
        for x in j:
            worksheet.write(row, column, x)
            print(x, row, column)
            column += 1
        row += 1

workbook.close()
print("done")
